from ast import Break
from asyncio import queues
from numbers import Number
import time
from tokenize import Name
import openpyxl
from openpyxl import load_workbook, Workbook
import selenium
import datetime
from selenium import webdriver
from selenium.webdriver import ActionChains

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException

#어떤 책 답변할 것인지
print("노베:35 시발점 수1: 83 수2: 78 미적(상):58 미적(하):57 기하:59 뉴런 수1:12 수2: 11 미적: 9")
Lec_num = int(input("어떤 거 답변할거니?: "))

#엑셀 이름 만드는데 쓸 변수
Lec_name = str(Lec_num)

#로그인 -> 질문 창 (아무것도 안하고 있으면 됨)
URL = 'https://tzone.megastudy.net/'

#창 열기
driver = webdriver.Chrome(executable_path='D:\자동 답변 프로그램\chromedriver.exe')
driver.get(url=URL)

driver.maximize_window()

#로그인
QA_buttion = driver.find_element_by_xpath('/html/body/form/div/div[1]/label[2]/input')
QA_buttion.click()

ID = driver.find_element_by_xpath('//*[@id="id"]')
ID.send_keys('heeseopark')

ID_send = driver.find_element_by_xpath('//*[@id="sp_smsct"]')
ID_send.click()

PW = driver.find_element_by_xpath('//*[@id="passwd"]')
PW.send_keys('heeseo1099')

Login_button = driver.find_element_by_xpath('//*[@id="sp_login"]')
Login_button.click()

driver.implicitly_wait(time_to_wait=5)

#대기 답변 넘어가기
Go_to_Answer_Tap = driver.find_element_by_xpath('//*[@id="a594"]')
Go_to_Answer_Tap.click()

Go_to_Waiting_Answer_Tap = driver.find_element_by_xpath('//*[@id="aa4233"]')
Go_to_Waiting_Answer_Tap.click()

driver.switch_to.frame('ifrmContents')

Lecture_Name = driver.find_element_by_xpath("/html/body/div[2]/form/label[1]/select")
Lecture_Name.click()

#강좌 option[n] 노베:35 시발점 수1: 83 수2: 78 미적(상):58 미적(하):57 기하:59 뉴런 수1:12 수2: 11 미적: 9
Lecture_Value = driver.find_element_by_xpath("/html/body/div[2]/form/label[1]/select/option[%d]" %Lec_num) 
Lecture_Value.click()

#엑셀 만들기
wb = openpyxl.Workbook()
ws = wb.active
write_wb = Workbook()
write_ws = write_wb.create_sheet('생성시트')

#엑셀 데이터 삽입, 저장 변수들
Q_Num_Xpath = list(range(20))
Q_Num = list(range(20))
Name_Xpath = list(range(20))
Name = list(range(20))
Book_Xpath = list(range(20))
Book = list(range(20))
Enter = list(range(20))
question_Xpath = list(range(20))
question = list(range(20))
new = list(range(20))
one_line = list()

#xpath 체크 함수
def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

#페이지 번호 존재 체크
for k in range(10):
    check_page = check_exists_by_xpath("/html/body/div[3]/div[3]/div/a[%d]" %k+1)
    if check_page == False:
        break



for i in range(1,21):
    check = check_exists_by_xpath("/html/body/div[3]/div[2]/table/tbody/tr[%d]/td[1]" %i) #문제 확인 함수
    if check == False:
        break
    Q_Num_Xpath[i-1] = driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/tbody/tr[%d]/td[1]" %i) #문제번호
    Q_Num[i-1] = Q_Num_Xpath[i-1].text

    ws.cell(row=1, column=i).value = Q_Num[i-1]

    Name_Xpath[i-1] = driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/tbody/tr[%d]/td[6]" %i)  #학생이름
    Name[i-1] = Name_Xpath[i-1].text
    ws.cell(row=2, column=i).value = Name[i-1]

    Book_Xpath[i-1] = driver.find_element_by_xpath("/html/body/div[3]/div[2]/table/tbody/tr[%d]/td[9]" %i)  #교재 구매 여부
    Book[i-1] = Book_Xpath[i-1].text
    ws.cell(row=3, column=i).value = Book[i-1]

    Enter = driver.find_element_by_xpath('/html/body/div[3]/div[2]/table/tbody/tr[%d]/td[5]/a' %i)  #질문 내용
    Enter.click()
    question_Xpath[i-1] = driver.find_element_by_xpath('/html/body/div[1]/table/tbody/tr[5]/td')
    question[i-1] = question_Xpath[i-1].text
    str_question = question[i-1]
    str(str_question).replace('=', '-')
    print(str_question)
    one_line = str_question.split(sep="\n")
    one_line_len = len(one_line)
    for k in range(one_line_len):
        ws.cell(row=4+k, column=i).value = one_line[k]
    driver.back()
    driver.switch_to.frame('ifrmContents')

#get number of question
num_of_question = 0
while True:
    if ws.cell(row=1, column=num_of_question+1).value != None:
        num_of_question = num_of_question + 1
    if ws.cell(row=1, column=num_of_question+1).value == None:
        num_of_question = 0
        break

#답변 금지 학생 목록
No_answer_students = ["정*성(royj****)", "정*이(jjkc****)", "황*하(king****)", "이*민(tkwk****)", "주*빈(seun****)", "정*희(jenn****)", "노*서 (nohm****)", "최*서 (sshs****)", "김*원(wldh****)", "박*영(ky65****)", "이*운(dong****)", "한*준 (ybs1****)", "서*림 (shr2****)", "정*희(jenn****)", "박*윤(jwdo****)", "김*준 (jun0****)", "장*진 (minj****)", "강*민(sumi****)", "정*원 (seon****)", "홍*호(gilh****)", "주*우 (best****)", "박*연(yeon****)", "김*원(jwpc****)", "권*영(nayo****)", "이*영(lyy1****)"]

#답변 안되는 것 삭제
def check_ans_available():
    for i in range(1, num_of_question+1):
        if ws.cell(row=3, column=num_of_question+1).value == "N": #교재 구매 안한 경우 열 삭제
            ws.delete_cols[num_of_question+1]
            num_of_question = num_of_question - 1


print("오늘 문제 가능이면 1, 불가능이면 0: ")
today_answer = int(input())

#엑셀 저장
s1 = time.strftime('%Y-%m-%d', time.localtime(time.time()))
'2018-05-19'
Excel_name = Lec_name + s1

wb.save('%s.xlsx' %Excel_name)